from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 영어 수학
#번호 (국어) 영어 수학 으로 국어를 추가하려면

ws.move_range("B1:C11",row=0, cols=1)
ws["B1"].value = "국어"  #b1셀에 국어 삽입

wb.save("sample_korean.xlsx")


