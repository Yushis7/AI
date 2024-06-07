from openpyxl import load_workbook
wb = load_workbook("sample_merge.xlsx")
ws = wb.active

ws.unmerge_cells('b2:d2') #b2부터d2까지 합체
# ws["b2"].value = "merged cell "

wb.save('sample_unmerge.xslx')