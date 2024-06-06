from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells('b2:d2') #b2부터d2까지 합체
ws["b2"].value = "merged cell "

wb.save('sample_merge.xslx')