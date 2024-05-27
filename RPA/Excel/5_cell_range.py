from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["번호","영어","수학"])
for i in range(1,11): # 10개 데이터 넣기
    ws.append([i,randint(0,100),randint(0,100)])

col_B = ws["B"] # 영어 column 만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] #영어,수학,colmun 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 1번쨰 row 만 가져오기
# for cell in row_title:
#     print(cell.value)

row_range= ws[2:6] # 1번쨰 title을 제외하고 2번쨰 줄에서 6번째 줄 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string

# row_rnage = ws[2:ws.max_row] # 2번쨰 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         print(cell.coordinate, end=" ") # A10,AZ250
#         XY = coordinate_from_string(cell.coordinate)
#         print(XY[0],end=" ") #A
#         print(XY[1],end=" ") # 1

        # print(cell.value, end = " ")
    # print()

#모든 rows 긁어오기
# print(tuple(ws.rows))

#모든 columns 긁어오기
# print(tuple(ws.columns))


#전체 row에서 몇행만 가져오기
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#    print(row[2].value)

#전체 columns에서 몇행만 가져오기
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#    print(column[0].value)

# 1번row부터 5번 row까지 2번col부터 3col까지 지정 출력하기
for row in ws.iter_rows(min_row=1,max_row=5,min_col=2,max_col=3): 
    print(row[0].value,row[1].value)


wb.save("sample1.xlsx")



1:02:24

