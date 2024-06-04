from openpyxl.styles import Font,Border,Side,PatternFill,alignment
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active


#번호,영어,수학
a1 = ws["A1"] #번호
b1 = ws["B1"] #영어
c1 = ws["C1"] #수학

#A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
#행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50
#스타일 고정
a1.font = Font(color="FF0000", italic=True,bold=True) #글자색 빨갛게 이태릭체 굵게
b1.font = Font(color="CC33FF", name="Arial",stirke=True) #폰트 arial 취소선
c1.font = Font(color="0000FF", size=20,underline=1) #글자크기 20 밑줄
#테두리 적용
thin_border = Border(left=Side(style="thin"))
a1.border=  thin_border

#90점 넘는 셀에 대해서 초록색으로 표시
for row in ws.rows:
    for cell in row:
        #각 셀에 대해서 정렬
        cell.alignment = alignment(horizontal = "center",vertical="center")
        if cell.column == 1: #A 번호열은 제외
            continue

        if isinstance(cell.value, int) and cell.value >90:
            cell.fill = PatternFill(fgColor="00FF00",fill_type="solid") #초록배경 솔리드 타입
            cell.font = Font(color="FF0000") #폰트 색상 변경

#틀 고정
ws.freeze_panes = "B2" #B2 기준으로 틀 고정


wb.save("sample_style.xlsx")
