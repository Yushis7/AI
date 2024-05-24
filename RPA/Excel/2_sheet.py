from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #시트 기본이름으로 새로 만들기 이름변경은()에 이름넣기
ws.title = "Mysheet" #sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" #RGB형태로 색상 변경

#Sheet,Mysheet,Yoursheet 
ws1 = wb.create_sheet("YourSheet") #주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) #2번쨰 인덱스에 시트 생성

new_ws = wb["NewSheet"] # Dict 형태로 sheet에 접근

print(wb.Sheetnames) # 모든 sheet이름 확인

#sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"  # copied sheet에도 a1에 test가 적힌다



wb.save("sample.xlsx")

